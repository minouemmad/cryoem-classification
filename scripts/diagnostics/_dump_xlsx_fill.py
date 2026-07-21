import openpyxl

wb = openpyxl.load_workbook(r'docs/GCER Population Tables.xlsx', data_only=True)
wbc = openpyxl.load_workbook(r'docs/GCER Population Tables.xlsx')  # styled
ws = wb['Current Processing']
wsc = wbc['Current Processing']

rows = list(ws.iter_rows(values_only=True))
hdr = [('' if v is None else str(v).strip()) for v in rows[0]]


def find(*keys):
    for i, h in enumerate(hdr):
        if all(k.lower() in (h or '').lower() for k in keys):
            return i
    return None


i_ds = find('Dataset', 'Location') or 0
i_cond = find('Construct')
i_prio = find('Priority')
i_npart = find('Particles')

for r in range(2, wsc.max_row + 1):
    dsx = wsc.cell(r, i_ds + 1)
    ds = dsx.value
    cond = wsc.cell(r, (i_cond or 0) + 1).value if i_cond is not None else ''
    if not (ds or cond):
        continue
    fill = dsx.fill
    rgb = ''
    try:
        if fill and fill.fgColor and fill.patternType:
            rgb = fill.fgColor.rgb or (fill.fgColor.theme, fill.fgColor.tint)
    except Exception:
        rgb = '?'
    prio = wsc.cell(r, (i_prio or 0) + 1).value if i_prio is not None else ''
    npart = wsc.cell(r, (i_npart or 0) + 1).value if i_npart is not None else ''
    print(f'row{r:>3} fill={str(rgb):>18} prio={str(prio):>5} n={str(npart):>10} '
          f'{str(ds)[:22]:<22} | {str(cond)}')
