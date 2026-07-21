import sys
import openpyxl

wb = openpyxl.load_workbook(r'docs/GCER Population Tables.xlsx', data_only=True)


def short(s, n=70):
    s = (s or '').replace('\n', ' ').replace('\r', ' ').strip()
    return (s[:n] + '...') if len(s) > n else s


def cols_of(sheet):
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [('' if v is None else str(v).strip()) for v in rows[0]]
    return hdr


def find(hdr, *keys):
    for i, h in enumerate(hdr):
        hl = (h or '').lower()
        if all(k.lower() in hl for k in keys):
            return i
    return None


def table(sheet, want):
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [('' if v is None else str(v).strip()) for v in rows[0]]
    idx = {}
    for label, keys in want.items():
        for keyset in keys:
            j = find(hdr, *keyset)
            if j is not None:
                idx[label] = j
                break
    print('\n===== SHEET:', sheet, '=====')
    print('columns used:', {k: hdr[v] for k, v in idx.items()})
    for r in rows[1:]:
        vals = [('' if v is None else str(v).strip()) for v in r]
        ds = vals[idx.get('ds', 0)] if 'ds' in idx else ''
        cond = vals[idx['cond']] if 'cond' in idx and idx['cond'] < len(vals) else ''
        if not (ds or cond):
            continue
        if not cond and not ds:
            continue
        parts = []
        for k in ('ds', 'cond', 'date', 'prio', 'npart', 'optpix', 'procpix', 'status'):
            if k in idx and idx[k] < len(vals):
                v = vals[idx[k]]
                if v:
                    parts.append(f'{k}={short(v, 55)}')
        line = ' | '.join(parts)
        if line.strip():
            print('-', line)


want = {
    'ds': [['Dataset', 'Location'], ['Dataset']],
    'cond': [['Construct'], ['Condition']],
    'date': [['Date']],
    'prio': [['Priority']],
    'npart': [['Particles'], ['hCFTR']],
    'optpix': [['Optimized', 'Pixel']],
    'procpix': [['Processing', 'Pixel']],
    'status': [['Status']],
}

for s in (sys.argv[1:] or ['Current Processing', 'Overview', 'Best Refinements']):
    table(s, want)
